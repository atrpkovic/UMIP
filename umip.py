#!/usr/bin/env python3
"""
UMIP - Universal Marketing Intelligence Platform
Single-file implementation for keyword trend analysis using Google Trends via SerpAPI.
"""

import os
import json
import hashlib
import requests
import pandas as pd
import numpy as np
from datetime import datetime, timedelta
from scipy import stats
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from dotenv import load_dotenv

# Configuration
CACHE_DIR = Path("cache")
OUTPUT_DIR = Path("output")
SERPAPI_BASE_URL = "https://serpapi.com/search"

# Trend classification thresholds
P_VALUE_THRESHOLD = 0.05  # Statistical significance threshold


def load_config():
    """Load environment variables."""
    load_dotenv()
    api_key = os.getenv("SERPAPI_KEY")
    if not api_key:
        raise ValueError("SERPAPI_KEY not found in environment. Create a .env file with SERPAPI_KEY=your_key")
    return {"serpapi_key": api_key}


def get_cache_path(keyword: str) -> Path:
    """Generate cache file path for a keyword."""
    CACHE_DIR.mkdir(exist_ok=True)
    hash_key = hashlib.md5(keyword.lower().encode()).hexdigest()[:12]
    return CACHE_DIR / f"trends_{hash_key}.json"


def fetch_google_trends(keyword: str, api_key: str, use_cache: bool = True) -> dict:
    """
    Fetch Google Trends data via SerpAPI.
    Returns 5 years of weekly data.
    """
    cache_path = get_cache_path(keyword)
    
    # Check cache first
    if use_cache and cache_path.exists():
        cache_age = datetime.now().timestamp() - cache_path.stat().st_mtime
        if cache_age < 86400 * 7:  # Cache valid for 7 days
            with open(cache_path, "r") as f:
                cached = json.load(f)
                print(f"  [CACHE] Using cached data for '{keyword}'")
                return cached
    
    # Fetch from SerpAPI
    print(f"  [API] Fetching Google Trends for '{keyword}'...")
    params = {
        "engine": "google_trends",
        "q": keyword,
        "date": "today 5-y",  # 5 years of data
        "api_key": api_key
    }
    
    try:
        response = requests.get(SERPAPI_BASE_URL, params=params, timeout=30)
        response.raise_for_status()
        data = response.json()
        
        # Cache the response
        with open(cache_path, "w") as f:
            json.dump(data, f)
        
        return data
    except requests.RequestException as e:
        print(f"  [ERROR] Failed to fetch trends for '{keyword}': {e}")
        return None


def parse_trends_data(raw_data: dict) -> pd.DataFrame:
    """
    Parse SerpAPI Google Trends response into a DataFrame.
    Returns DataFrame with columns: date, interest, week_num
    """
    if not raw_data:
        return pd.DataFrame()
    
    # SerpAPI returns interest_over_time with timeline_data
    timeline = raw_data.get("interest_over_time", {}).get("timeline_data", [])
    
    if not timeline:
        return pd.DataFrame()
    
    records = []
    for i, point in enumerate(timeline):
        # Use timestamp field (Unix timestamp) - much more reliable than parsing date string
        timestamp = point.get("timestamp", "")
        if not timestamp:
            continue
        
        dt = datetime.fromtimestamp(int(timestamp))
        
        # Parse interest value
        values = point.get("values", [])
        interest = values[0].get("extracted_value", 0) if values else 0
        
        records.append({
            "date": dt,
            "interest": interest,
            "week_num": i + 1
        })
    
    df = pd.DataFrame(records)
    
    if df.empty:
        return df
    
    df = df.sort_values("date").reset_index(drop=True)
    df["week_num"] = range(1, len(df) + 1)
    
    return df


def calculate_trend_metrics(df: pd.DataFrame) -> dict:
    """
    Calculate linear regression and trend metrics.
    Returns dict with slope, p_value, annual_growth, trend_classification.
    """
    if df.empty or len(df) < 10:
        return {
            "slope": None,
            "p_value": None,
            "annual_growth": None,
            "trend": "Insufficient Data"
        }
    
    x = df["week_num"].values
    y = df["interest"].values
    
    # Linear regression
    slope, intercept, r_value, p_value, std_err = stats.linregress(x, y)
    
    # Calculate estimated annual growth (52 weeks * slope)
    annual_growth = slope * 52
    
    # Classify trend
    if p_value > P_VALUE_THRESHOLD:
        trend = "Not Significant"
    elif slope > 0:
        trend = "Growth"
    else:
        trend = "Decline"
    
    return {
        "slope": slope,
        "p_value": p_value,
        "annual_growth": annual_growth,
        "trend": trend
    }


def calculate_seasonality(df: pd.DataFrame) -> dict:
    """
    Calculate quarterly seasonality metrics.
    Returns dict with Q1-Q4 averages and peak quarter.
    """
    if df.empty:
        return {"Q1": None, "Q2": None, "Q3": None, "Q4": None, "peak_quarter": None}
    
    df = df.copy()
    df["quarter"] = df["date"].dt.quarter
    
    quarterly_avg = df.groupby("quarter")["interest"].mean()
    
    q_values = {
        "Q1": quarterly_avg.get(1, 0),
        "Q2": quarterly_avg.get(2, 0),
        "Q3": quarterly_avg.get(3, 0),
        "Q4": quarterly_avg.get(4, 0)
    }
    
    peak = max(q_values, key=q_values.get)
    q_values["peak_quarter"] = peak
    
    return q_values


def calculate_monthly_seasonality(df: pd.DataFrame) -> dict:
    """
    Calculate monthly seasonality metrics.
    Returns dict with Jan-Dec averages and peak month.
    """
    if df.empty:
        return {m: None for m in ["Jan", "Feb", "Mar", "Apr", "May", "Jun", 
                                   "Jul", "Aug", "Sep", "Oct", "Nov", "Dec", "peak_month"]}
    
    df = df.copy()
    df["month"] = df["date"].dt.month
    
    monthly_avg = df.groupby("month")["interest"].mean()
    
    month_names = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", 
                   "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
    
    m_values = {}
    for i, name in enumerate(month_names, 1):
        m_values[name] = monthly_avg.get(i, 0)
    
    peak = max(month_names, key=lambda x: m_values[x])
    m_values["peak_month"] = peak
    
    return m_values


def analyze_keyword(keyword: str, api_key: str) -> dict:
    """
    Full analysis pipeline for a single keyword.
    """
    print(f"\nAnalyzing: {keyword}")
    
    # Fetch data
    raw_data = fetch_google_trends(keyword, api_key)
    
    # Parse
    df = parse_trends_data(raw_data)
    
    if df.empty:
        print(f"  [WARN] No data available for '{keyword}'")
        return {
            "keyword": keyword,
            "trend_metrics": {"slope": None, "p_value": None, "annual_growth": None, "trend": "No Data"},
            "seasonality": {"Q1": None, "Q2": None, "Q3": None, "Q4": None, "peak_quarter": None},
            "monthly_seasonality": {m: None for m in ["Jan", "Feb", "Mar", "Apr", "May", "Jun", 
                                                       "Jul", "Aug", "Sep", "Oct", "Nov", "Dec", "peak_month"]},
            "raw_data": pd.DataFrame()
        }
    
    # Calculate metrics
    trend_metrics = calculate_trend_metrics(df)
    seasonality = calculate_seasonality(df)
    monthly_seasonality = calculate_monthly_seasonality(df)
    
    p_val = trend_metrics['p_value']
    print(f"  Trend: {trend_metrics['trend']} | p-value: {p_val}")
    print(f"  Peak Quarter: {seasonality['peak_quarter']} | Peak Month: {monthly_seasonality['peak_month']}")
    
    return {
        "keyword": keyword,
        "trend_metrics": trend_metrics,
        "seasonality": seasonality,
        "monthly_seasonality": monthly_seasonality,
        "raw_data": df
    }


def create_output_report(results: list, output_path: Path):
    """
    Generate Excel report with trend analysis, seasonality, and raw data.
    """
    wb = Workbook()
    
    # ===== Sheet 1: Trend Analysis =====
    ws_trend = wb.active
    ws_trend.title = "Trend Analysis"
    
    # Headers
    headers = ["KW (Category)", "Slope (m)", "p-value", "Est. Annual Growth", "Trend", "Q1", "Q2", "Q3", "Q4", "Peak Quarter"]
    ws_trend.append(headers)
    
    # Style headers
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for col, cell in enumerate(ws_trend[1], 1):
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    
    # Data rows
    trend_colors = {
        "Growth": "C6EFCE",      # Green
        "Decline": "FFC7CE",     # Red
        "Not Significant": "FFEB9C",  # Yellow
        "No Data": "D9D9D9",     # Gray
        "Insufficient Data": "D9D9D9"
    }
    
    for r in results:
        tm = r["trend_metrics"]
        s = r["seasonality"]
        row = [
            r["keyword"],
            tm["slope"],
            tm["p_value"],
            tm["annual_growth"],
            tm["trend"],
            s["Q1"],
            s["Q2"],
            s["Q3"],
            s["Q4"],
            s["peak_quarter"]
        ]
        ws_trend.append(row)
        
        # Color code trend column
        row_num = ws_trend.max_row
        trend_cell = ws_trend.cell(row=row_num, column=5)
        fill_color = trend_colors.get(tm["trend"], "FFFFFF")
        trend_cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
    
    # Adjust column widths
    ws_trend.column_dimensions["A"].width = 25
    for col in ["B", "C", "D", "E", "F", "G", "H", "I", "J"]:
        ws_trend.column_dimensions[col].width = 15
    
    # ===== Sheet 2: Seasonality Analysis =====
    ws_season = wb.create_sheet("Seasonality Analysis")
    
    season_headers = ["KW (Category)", "Q1", "Q2", "Q3", "Q4", "Peak Quarter"]
    ws_season.append(season_headers)
    
    for cell in ws_season[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    
    for r in results:
        s = r["seasonality"]
        row = [r["keyword"], s["Q1"], s["Q2"], s["Q3"], s["Q4"], s["peak_quarter"]]
        ws_season.append(row)
        
        # Highlight peak quarter
        row_num = ws_season.max_row
        peak = s["peak_quarter"]
        if peak:
            q_col_map = {"Q1": 2, "Q2": 3, "Q3": 4, "Q4": 5}
            peak_col = q_col_map.get(peak)
            if peak_col:
                ws_season.cell(row=row_num, column=peak_col).fill = PatternFill(
                    start_color="92D050", end_color="92D050", fill_type="solid"
                )
    
    ws_season.column_dimensions["A"].width = 25
    for col in ["B", "C", "D", "E", "F"]:
        ws_season.column_dimensions[col].width = 15
    
    # ===== Sheet 3: Monthly Seasonality Analysis =====
    ws_monthly = wb.create_sheet("Monthly Seasonality")
    
    month_names = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", 
                   "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
    monthly_headers = ["KW (Category)"] + month_names + ["Peak Month"]
    ws_monthly.append(monthly_headers)
    
    for cell in ws_monthly[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    
    for r in results:
        ms = r["monthly_seasonality"]
        row = [r["keyword"]] + [ms[m] for m in month_names] + [ms["peak_month"]]
        ws_monthly.append(row)
        
        # Highlight peak month
        row_num = ws_monthly.max_row
        peak = ms["peak_month"]
        if peak:
            peak_col = month_names.index(peak) + 2  # +2 because col 1 is keyword
            ws_monthly.cell(row=row_num, column=peak_col).fill = PatternFill(
                start_color="92D050", end_color="92D050", fill_type="solid"
            )
    
    ws_monthly.column_dimensions["A"].width = 25
    for col in ["B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M", "N"]:
        ws_monthly.column_dimensions[col].width = 8
    
    # ===== Sheet 4: Chart Examples (Raw Data) =====
    ws_charts = wb.create_sheet("Chart Examples")
    
    current_row = 1
    for r in results:
        df = r["raw_data"]
        if df.empty:
            continue
        
        # Section header
        ws_charts.cell(row=current_row, column=1, 
                      value=f"Example for category: {r['trend_metrics']['trend']} (Term: {r['keyword']})")
        ws_charts.cell(row=current_row, column=1).font = Font(bold=True)
        current_row += 1
        
        # Column headers
        ws_charts.cell(row=current_row, column=1, value="Date")
        ws_charts.cell(row=current_row, column=2, value="Interest")
        ws_charts.cell(row=current_row, column=3, value="Week")
        current_row += 1
        
        # Data
        for _, data_row in df.iterrows():
            ws_charts.cell(row=current_row, column=1, value=data_row["date"].strftime("%Y-%m-%d"))
            ws_charts.cell(row=current_row, column=2, value=data_row["interest"])
            ws_charts.cell(row=current_row, column=3, value=data_row["week_num"])
            current_row += 1
        
        current_row += 3  # Gap between keywords
    
    ws_charts.column_dimensions["A"].width = 20
    ws_charts.column_dimensions["B"].width = 12
    ws_charts.column_dimensions["C"].width = 10
    
    # Save
    OUTPUT_DIR.mkdir(exist_ok=True)
    wb.save(output_path)
    print(f"\n✓ Report saved to: {output_path}")


def load_keywords(filepath: str) -> list:
    """Load keywords from CSV file."""
    df = pd.read_csv(filepath, encoding="utf-8-sig")
    # Handle BOM and whitespace
    if "keyword" in df.columns:
        return df["keyword"].str.strip().tolist()
    # Fallback to first column
    return df.iloc[:, 0].str.strip().tolist()


def main():
    """Main entry point."""
    print("=" * 60)
    print("UMIP - Universal Marketing Intelligence Platform")
    print("=" * 60)
    
    # Load config
    try:
        config = load_config()
    except ValueError as e:
        print(f"\n[ERROR] {e}")
        return
    
    # Load keywords
    keywords_file = "keywords_master.csv"
    if not Path(keywords_file).exists():
        print(f"\n[ERROR] Keywords file not found: {keywords_file}")
        print("Create a CSV file with a 'keyword' column.")
        return
    
    keywords = load_keywords(keywords_file)
    print(f"\nLoaded {len(keywords)} keywords: {keywords}")
    
    # Analyze each keyword
    results = []
    for kw in keywords:
        result = analyze_keyword(kw, config["serpapi_key"])
        results.append(result)
    
    # Generate report
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = OUTPUT_DIR / f"umip_analysis_{timestamp}.xlsx"
    create_output_report(results, output_path)
    
    # Summary
    print("\n" + "=" * 60)
    print("ANALYSIS SUMMARY")
    print("=" * 60)
    
    growth_count = sum(1 for r in results if r["trend_metrics"]["trend"] == "Growth")
    decline_count = sum(1 for r in results if r["trend_metrics"]["trend"] == "Decline")
    ns_count = sum(1 for r in results if r["trend_metrics"]["trend"] == "Not Significant")
    
    print(f"  Growth trends:         {growth_count}")
    print(f"  Decline trends:        {decline_count}")
    print(f"  Not Significant:       {ns_count}")
    print(f"\nOutput: {output_path}")


if __name__ == "__main__":
    main()